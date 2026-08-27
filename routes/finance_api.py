import io
from datetime import datetime

import pandas as pd
from db import ObjectId
from flask import jsonify, make_response, request, send_file, session
from utils import patient_financial_summary
from utils import safe_int_amount as _safe_int_amount


def _object_id_or_none(raw_id):
    try:
        return ObjectId(str(raw_id))
    except Exception:
        return None


def _extract_payment_patient_name(note=''):
    note = str(note or '')
    if 'Partial payment from ' in note:
        return note.split('Partial payment from ')[1].split(' via ')[0].strip()
    if 'Payment from ' in note:
        return note.split('Payment from ')[1].split(' via ')[0].strip()
    if 'Initial Advance from ' in note:
        return note.split('Initial Advance from ')[1].split(' (')[0].strip()
    return 'Unknown'


def _build_payment_record_note(existing_note, patient_name, payment_method):
    patient_name = patient_name or 'Unknown'
    existing_note = str(existing_note or '')
    large_amount_suffix = ' (Large Amount)' if '(Large Amount)' in existing_note else ''

    if existing_note.startswith('Initial Advance from '):
        return f"Initial Advance from {patient_name} (Admission)"
    if existing_note.startswith('Partial payment from '):
        return f"Partial payment from {patient_name} via {payment_method}{large_amount_suffix}"
    return f"Payment from {patient_name} via {payment_method}{large_amount_suffix}"


def _serialize_manual_receipt(doc):
    return {
        'id': str(doc.get('_id')),
        'patient_id': str(doc.get('patient_id')) if doc.get('patient_id') else '',
        'patient_name': doc.get('patient_name', ''),
        'father_name': doc.get('father_name', ''),
        'age': doc.get('age', ''),
        'cnic': doc.get('cnic', ''),
        'contact_no': doc.get('contact_no', ''),
        'area': doc.get('area', ''),
        'address': doc.get('address', ''),
        'admission_date': doc.get('admission_date', ''),
        'discharge_date': doc.get('discharge_date', ''),
        'stay_days': doc.get('stay_days', 0),
        'monthly_fee': doc.get('monthly_fee', 0),
        'fee_amount': doc.get('fee_amount', 0),
        'rehab_next_month_amount': doc.get('rehab_next_month_amount', 0),
        'test_amount': doc.get('test_amount', 0),
        'canteen_amount': doc.get('canteen_amount', 0),
        'laundry_amount': doc.get('laundry_amount', 0),
        'barbar_amount': doc.get('barbar_amount', 0),
        'medicine_amount': doc.get('medicine_amount', 0),
        'other_amount': doc.get('other_amount', 0),
        'received_amount': doc.get('received_amount', 0),
        'net_balance': doc.get('net_balance', 0),
        'notes': doc.get('notes', ''),
        'created_by': doc.get('created_by', ''),
        'updated_by': doc.get('updated_by', ''),
        'created_at': doc.get('created_at').isoformat() if doc.get('created_at') else '',
        'updated_at': doc.get('updated_at').isoformat() if doc.get('updated_at') else ''
    }


def _month_start_n_months_ago(months_ago):
    today = datetime.now()
    target_month = today.month - months_ago
    target_year = today.year
    while target_month <= 0:
        target_month += 12
        target_year -= 1
    return datetime(target_year, target_month, 1)


def register_finance_api_routes(
    app,
    mongo,
    check_db,
    clean_input_data,
    decrypt_data,
    login_required,
    role_required,
    calculate_prorated_fee,
    get_current_user_id,
):
    def adjust_patient_received_amount(patient_id, delta_amount):
        patient_oid = _object_id_or_none(patient_id)
        if not patient_oid:
            return

        patient = mongo.db.patients.find_one({'_id': patient_oid}, {'receivedAmount': 1})
        if not patient:
            return

        current_received = _safe_int_amount(patient.get('receivedAmount', 0))
        new_total = max(current_received + _safe_int_amount(delta_amount), 0)
        mongo.db.patients.update_one(
            {'_id': patient_oid},
            {'$set': {'receivedAmount': str(new_total)}}
        )

    @app.route('/api/expenses', methods=['GET'])
    @login_required
    def list_expenses():
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            cursor = mongo.db.expenses.find({'deleted_at': {'$exists': False}}).sort('date', -1)
            expenses = []
            for expense in cursor:
                expenses.append({
                    'id': str(expense.get('_id')),
                    'type': expense.get('type', 'outgoing'),
                    'amount': expense.get('amount', 0),
                    'category': expense.get('category', ''),
                    'note': expense.get('note', ''),
                    'date': expense.get('date').isoformat() if expense.get('date') else '',
                    'recorded_by': expense.get('recorded_by', ''),
                    'auto': False
                })

            try:
                patients = mongo.db.patients.find()
                total_fees = 0
                for patient in patients:
                    try:
                        total_fees += int(str(patient.get('monthlyFee', '0')).replace(',', ''))
                    except ValueError:
                        pass

                today = datetime.now()
                start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                pipeline = [
                    {'$match': {'date': {'$gte': start_of_month}}},
                    {'$group': {'_id': None, 'total_sales': {'$sum': '$amount'}}}
                ]
                sales_result = list(mongo.db.canteen_sales.aggregate(pipeline))
                total_canteen = sales_result[0]['total_sales'] if sales_result else 0

                today_iso = datetime.now().date().isoformat()
                expenses.insert(0, {
                    'id': 'auto-canteen',
                    'type': 'incoming',
                    'amount': total_canteen,
                    'category': 'Canteen Sales (auto)',
                    'note': 'Automatically calculated from canteen sales this month',
                    'date': today_iso,
                    'recorded_by': 'system',
                    'auto': True
                })
                expenses.insert(0, {
                    'id': 'auto-fees',
                    'type': 'incoming',
                    'amount': total_fees,
                    'category': 'Monthly Fees (auto)',
                    'note': 'Automatically calculated from patient monthly fees',
                    'date': today_iso,
                    'recorded_by': 'system',
                    'auto': True
                })
            except Exception as error:
                print(f"Auto income calc error: {error}")

            return jsonify(expenses)
        except Exception as error:
            print(f"Expenses list error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/expenses', methods=['POST'])
    @role_required(['Admin'])
    def add_expense():
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        data = clean_input_data(request.json or {})
        required = ['type', 'amount', 'category']
        if not all(key in data for key in required):
            return jsonify({"error": "Missing fields"}), 400

        try:
            amount = int(str(data.get('amount', 0)).replace(',', ''))
        except ValueError:
            return jsonify({"error": "Amount must be a number"}), 400

        expense = {
            'type': data.get('type', 'outgoing'),
            'amount': amount,
            'category': data.get('category', ''),
            'note': data.get('note', ''),
            'date': datetime.fromisoformat(data.get('date')) if data.get('date') else datetime.now(),
            'recorded_by': session.get('username', 'System'),
            'created_at': datetime.now()
        }
        try:
            result = mongo.db.expenses.insert_one(expense)
            return jsonify({"message": "Expense saved", "id": str(result.inserted_id)}), 201
        except Exception as error:
            print(f"Add expense error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/expenses/<id>', methods=['DELETE'])
    @role_required(['Admin'])
    def delete_expense(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            result = mongo.db.expenses.update_one({'_id': ObjectId(id)}, {'$set': {'deleted_at': datetime.now()}})
            if result.modified_count:
                return jsonify({"message": "Expense soft-deleted"})
            return jsonify({"error": "Expense not found"}), 404
        except Exception as error:
            print(f"Delete expense error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/expenses/summary', methods=['GET'])
    @login_required
    def expenses_summary():
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        today = datetime.now()
        start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        try:
            pipeline = [
                {'$match': {'date': {'$gte': start_of_month}, 'deleted_at': {'$exists': False}}},
                {'$group': {'_id': '$type', 'total': {'$sum': '$amount'}}}
            ]
            summary_data = list(mongo.db.expenses.aggregate(pipeline))
            incoming = 0
            outgoing = 0
            for item in summary_data:
                if item['_id'] == 'incoming':
                    incoming = item['total']
                elif item['_id'] == 'outgoing':
                    outgoing = item['total']

            return jsonify({
                'incoming': incoming,
                'outgoing': outgoing,
                'net': incoming - outgoing
            })
        except Exception as error:
            print(f"Expenses summary error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/accounts/summary', methods=['GET'])
    @role_required(['Admin'])
    def get_accounts_summary():
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            patients = list(mongo.db.patients.find({}, {
                'name': 1, 'fatherName': 1, 'admissionDate': 1,
                'monthlyFee': 1, 'address': 1, 'age': 1,
                'laundryStatus': 1, 'laundryAmount': 1, 'receivedAmount': 1,
                'isDischarged': 1
            }))

            pipeline = [{'$group': {'_id': '$patient_id', 'total_sales': {'$sum': '$amount'}}}]
            sales_data = list(mongo.db.canteen_sales.aggregate(pipeline))
            sales_map = {str(item['_id']): item['total_sales'] for item in sales_data}

            summary = []
            for patient in patients:
                patient_id = str(patient['_id'])
                admission_date = patient.get('admissionDate')
                days_elapsed = 0
                if admission_date:
                    try:
                        if isinstance(admission_date, str):
                            admission_dt = datetime.fromisoformat(admission_date.replace('Z', '+00:00'))
                        else:
                            admission_dt = admission_date
                        days_diff = (datetime.now() - admission_dt).days
                        days_elapsed = max(0, days_diff)
                    except Exception:
                        days_elapsed = 0

                monthly_fee = patient.get('monthlyFee', '0')
                calculated_fee = calculate_prorated_fee(monthly_fee, days_elapsed)

                summary.append({
                    'id': patient_id,
                    'name': decrypt_data(patient.get('name', '')),
                    'fatherName': patient.get('fatherName', ''),
                    'age': patient.get('age', ''),
                    'area': patient.get('address', ''),
                    'admissionDate': patient.get('admissionDate', ''),
                    'monthlyFee': monthly_fee,
                    'calculatedFee': calculated_fee,
                    'daysElapsed': days_elapsed,
                    'canteenTotal': sales_map.get(patient_id, 0),
                    'laundryStatus': patient.get('laundryStatus', False),
                    'laundryAmount': patient.get('laundryAmount', 0),
                    'receivedAmount': patient.get('receivedAmount', '0'),
                    'isDischarged': patient.get('isDischarged', False)
                })

            return jsonify(summary)
        except Exception as error:
            return jsonify({"error": str(error)}), 500

    @app.route('/api/payment-records', methods=['GET'])
    @role_required(['Admin'])
    def get_payment_records():
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            payments = list(mongo.db.expenses.find({
                'type': 'incoming',
                'category': 'Patient Fee'
            }).sort('date', -1))

            patient_ids = list(set(payment.get('patient_id') for payment in payments if payment.get('patient_id')))
            patient_map = {}
            if patient_ids:
                valid_ids = []
                for patient_id in patient_ids:
                    object_id = _object_id_or_none(patient_id)
                    if object_id:
                        valid_ids.append(object_id)
                patients = list(mongo.db.patients.find({'_id': {'$in': valid_ids}}, {'name': 1}))
                patient_map = {str(patient['_id']): decrypt_data(patient.get('name', 'Unknown')) for patient in patients}

            records = []
            for payment in payments:
                patient_id = payment.get('patient_id')
                patient_name = patient_map.get(str(patient_id)) if patient_id else None
                if not patient_name:
                    patient_name = _extract_payment_patient_name(payment.get('note', ''))

                records.append({
                    '_id': str(payment['_id']),
                    'patient_id': str(patient_id) if patient_id else '',
                    'patient_name': patient_name,
                    'amount': payment.get('amount', 0),
                    'date': payment.get('date').strftime('%Y-%m-%d') if payment.get('date') else 'N/A',
                    'payment_method': payment.get('payment_method', 'Cash'),
                    'recorded_by': payment.get('recorded_by', 'Admin'),
                    'screenshot': payment.get('screenshot', ''),
                    'note': payment.get('note', '')
                })

            return jsonify(records)
        except Exception as error:
            print(f"Payment Records Error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/payment-records/<id>', methods=['PUT'])
    @role_required(['Admin'])
    def update_payment_record(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        payment_oid = _object_id_or_none(id)
        if not payment_oid:
            return jsonify({"error": "Invalid payment record id"}), 400

        data = clean_input_data(request.json or {})
        amount = _safe_int_amount(data.get('amount'))
        if amount <= 0:
            return jsonify({"error": "Amount must be greater than zero"}), 400

        raw_date = str(data.get('date', '')).strip()
        try:
            payment_date = datetime.fromisoformat(raw_date) if raw_date else None
        except ValueError:
            return jsonify({"error": "Invalid payment date"}), 400

        try:
            payment_doc = mongo.db.expenses.find_one({
                '_id': payment_oid,
                'type': 'incoming',
                'category': 'Patient Fee'
            })
            if not payment_doc:
                return jsonify({"error": "Payment record not found"}), 404

            payment_method = data.get('payment_method') or payment_doc.get('payment_method', 'Cash')
            is_online_payment = str(payment_method).lower().startswith('online')
            screenshot = data.get('screenshot', payment_doc.get('screenshot', '')) if is_online_payment else ''

            patient_id = payment_doc.get('patient_id')
            patient_name = _extract_payment_patient_name(payment_doc.get('note', ''))
            patient_oid = _object_id_or_none(patient_id)
            if patient_oid:
                patient = mongo.db.patients.find_one({'_id': patient_oid}, {'name': 1})
                if patient and patient.get('name'):
                    patient_name = patient.get('name')

            old_amount = _safe_int_amount(payment_doc.get('amount'))
            amount_delta = amount - old_amount
            if amount_delta:
                adjust_patient_received_amount(patient_id, amount_delta)

            mongo.db.expenses.update_one(
                {'_id': payment_oid},
                {
                    '$set': {
                        'amount': amount,
                        'payment_method': payment_method,
                        'screenshot': screenshot,
                        'date': payment_date or payment_doc.get('date') or datetime.now(),
                        'note': _build_payment_record_note(payment_doc.get('note', ''), patient_name, payment_method),
                        'updated_at': datetime.now()
                    }
                }
            )
            return jsonify({"message": "Payment record updated"})
        except Exception as error:
            print(f"Payment Record Update Error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/payment-records/<id>', methods=['DELETE'])
    @role_required(['Admin'])
    def delete_payment_record(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        payment_oid = _object_id_or_none(id)
        if not payment_oid:
            return jsonify({"error": "Invalid payment record id"}), 400

        try:
            payment_doc = mongo.db.expenses.find_one({
                '_id': payment_oid,
                'type': 'incoming',
                'category': 'Patient Fee'
            })
            if not payment_doc:
                return jsonify({"error": "Payment record not found"}), 404

            adjust_patient_received_amount(
                payment_doc.get('patient_id'),
                -_safe_int_amount(payment_doc.get('amount'))
            )

            mongo.db.expenses.delete_one({'_id': payment_oid})
            return jsonify({"message": "Payment record deleted"})
        except Exception as error:
            print(f"Payment Record Delete Error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/finance/summary/<int:month>/<int:year>', methods=['GET'])
    @role_required(['Admin'])
    def get_finance_summary(month, year):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            month_start = datetime(year, month, 1).date()
            if month == 12:
                month_end = datetime(year + 1, 1, 1).date()
            else:
                month_end = datetime(year, month + 1, 1).date()

            def parse_date_only(raw_val):
                if not raw_val:
                    return None
                if isinstance(raw_val, datetime):
                    return raw_val.date()
                try:
                    return datetime.fromisoformat(str(raw_val).replace('Z', '+00:00')).date()
                except Exception:
                    try:
                        return datetime.strptime(str(raw_val)[:10], '%Y-%m-%d').date()
                    except Exception:
                        return None

            patients = list(mongo.db.patients.find({}, {
                'monthlyFee': 1,
                'admissionDate': 1,
                'isDischarged': 1,
                'dischargeDate': 1
            }))

            total_expected_incoming = 0
            for patient in patients:
                admission_date = parse_date_only(patient.get('admissionDate'))
                discharge_date = parse_date_only(patient.get('dischargeDate')) if patient.get('isDischarged') else None
                if admission_date and admission_date >= month_end:
                    continue
                if discharge_date and discharge_date < month_start:
                    continue
                total_expected_incoming += _safe_int_amount(patient.get('monthlyFee', 0))

            employees = list(mongo.db.employees.find())
            total_salaries = 0
            for employee in employees:
                try:
                    total_salaries += int(str(employee.get('pay', '0')).replace(',', ''))
                except Exception:
                    pass

            search_pattern = f"{year}-{month:02d}-"
            bills = list(mongo.db.utility_bills.find({'due_date': {'$regex': f'^{search_pattern}'}}))
            total_bills = sum(bill.get('amount', 0) for bill in bills)

            overheads = list(mongo.db.overheads.find({'month': month, 'year': year}))
            total_kitchen = 0
            total_others = 0
            total_pay_advance = 0
            for entry in overheads:
                total_kitchen += entry.get('kitchen', 0)
                total_others += entry.get('others', 0)
                total_pay_advance += entry.get('pay_advance', 0)

            start_date = datetime(year, month, 1)
            if month == 12:
                end_date = datetime(year + 1, 1, 1)
            else:
                end_date = datetime(year, month + 1, 1)

            canteen_agg = mongo.db.canteen_sales.aggregate([
                {'$match': {'date': {'$gte': start_date, '$lt': end_date}}},
                {'$group': {'_id': None, 'total': {'$sum': '$amount'}}}
            ])
            canteen_res = list(canteen_agg)
            total_canteen_auto = canteen_res[0]['total'] if canteen_res else 0

            total_estimated_overheads = total_salaries + total_bills + total_kitchen + total_others + total_pay_advance + total_canteen_auto

            return jsonify({
                'month': month,
                'year': year,
                'totalSalaries': total_salaries,
                'totalUtilityBills': total_bills,
                'totalKitchen': total_kitchen,
                'totalCanteenAuto': total_canteen_auto,
                'totalOthers': total_others,
                'totalPayAdvance': total_pay_advance,
                'totalEstimatedOverheads': total_estimated_overheads,
                'totalIncome': total_expected_incoming,
                'profit': total_expected_incoming - total_estimated_overheads
            })
        except Exception as error:
            print(f"Finance Summary Error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/payment-records/export', methods=['GET'])
    @role_required(['Admin'])
    def export_payment_records():
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        range_key = request.args.get('range', 'current')
        today = datetime.now()
        if range_key == 'six_months':
            start_date = _month_start_n_months_ago(5)
        else:
            start_date = datetime(today.year, today.month, 1)

        if today.month == 12:
            end_date = datetime(today.year + 1, 1, 1)
        else:
            end_date = datetime(today.year, today.month + 1, 1)

        try:
            payments = list(mongo.db.expenses.find({
                'type': 'incoming',
                'category': 'Patient Fee',
                'date': {'$gte': start_date, '$lt': end_date}
            }).sort('date', 1))

            rows = []

            def to_date(dt_val):
                if not dt_val:
                    return ''
                if isinstance(dt_val, datetime):
                    return dt_val
                try:
                    return datetime.fromisoformat(str(dt_val))
                except Exception:
                    return None

            patient_ids = list(set(payment.get('patient_id') for payment in payments if payment.get('patient_id')))
            patient_map = {}
            if patient_ids:
                valid_ids = []
                for patient_id in patient_ids:
                    object_id = _object_id_or_none(patient_id)
                    if object_id:
                        valid_ids.append(object_id)
                patients = list(mongo.db.patients.find({'_id': {'$in': valid_ids}}, {'name': 1}))
                patient_map = {str(patient['_id']): decrypt_data(patient.get('name', 'Unknown')) for patient in patients}

            for payment in payments:
                patient_id = payment.get('patient_id')
                note = payment.get('note', '')
                patient_name = patient_map.get(str(patient_id)) if patient_id else None
                if not patient_name:
                    patient_name = _extract_payment_patient_name(note)

                dt_val = to_date(payment.get('date'))
                rows.append({
                    'Patient Name': patient_name,
                    'Amount (PKR)': payment.get('amount', 0),
                    'Date': dt_val.strftime('%Y-%m-%d') if dt_val else '',
                    'Payment Mode': payment.get('payment_method', 'Cash'),
                    'Recorded By': payment.get('recorded_by', 'Admin'),
                    'Note': note
                })

            df = pd.DataFrame(rows)
            if df.empty:
                df = pd.DataFrame([{'Message': 'No payment records for selected range'}])

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Payment Records')
                worksheet = writer.sheets['Payment Records']
                worksheet.page_setup.paperSize = 9
                worksheet.page_setup.orientation = 'portrait'
                worksheet.page_setup.fitToWidth = 1
                worksheet.page_setup.fitToHeight = 0
                worksheet.print_options.horizontalCentered = True
                worksheet.page_margins.left = 0.75
                worksheet.page_margins.right = 0.75
                worksheet.page_margins.top = 0.75
                worksheet.page_margins.bottom = 0.75

            output.seek(0)
            filename = f"payment_records_{'six_months' if range_key == 'six_months' else 'current_month'}.xlsx"
            return send_file(
                output,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as error:
            print(f"Payment Records Export Error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/patients/<id>/payment', methods=['POST'])
    @role_required(['Admin'])
    def add_patient_payment(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            data = clean_input_data(request.json or {})
            amount_paid = _safe_int_amount(data.get('amount'))
            if amount_paid <= 0:
                return jsonify({"error": "Amount must be greater than zero"}), 400

            payment_method = data.get('payment_method', 'Cash')
            screenshot = data.get('screenshot', '') if str(payment_method).lower().startswith('online') else ''
            raw_payment_date = str(data.get('payment_date', '')).strip()
            try:
                payment_date = datetime.fromisoformat(raw_payment_date) if raw_payment_date else datetime.now()
            except ValueError:
                return jsonify({"error": "Invalid payment date"}), 400

            patient = mongo.db.patients.find_one({'_id': ObjectId(id)})
            if not patient:
                return jsonify({"error": "Patient not found"}), 404

            current_received_str = str(patient.get('receivedAmount', '0')).replace(',', '')
            try:
                current_received = int(current_received_str)
            except ValueError:
                current_received = 0

            new_total = current_received + amount_paid

            mongo.db.patients.update_one(
                {'_id': ObjectId(id)},
                {'$set': {'receivedAmount': str(new_total)}}
            )

            is_overpayment = False
            try:
                if amount_paid > 100000:
                    is_overpayment = True
            except Exception:
                pass

            expense_note = f"Payment from {patient.get('name')} via {payment_method}"
            if is_overpayment:
                expense_note += " (Large Amount)"

            mongo.db.expenses.insert_one({
                'type': 'incoming',
                'amount': amount_paid,
                'category': 'Patient Fee',
                'note': expense_note,
                'payment_method': payment_method,
                'patient_id': str(id),
                'screenshot': screenshot,
                'date': payment_date,
                'recorded_by': session.get('username', 'Admin'),
                'auto': True
            })

            return jsonify({"message": "Payment recorded successfully", "new_total": new_total})
        except Exception as error:
            print(f"Payment Error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/patients/<id>/discharge-bill', methods=['GET'])
    @role_required(['Admin', 'Doctor'])
    def generate_discharge_bill(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

            patient = mongo.db.patients.find_one({'_id': ObjectId(id)})
            if not patient:
                return jsonify({"error": "Patient not found"}), 404

            admission_date = patient.get('admissionDate')
            days_elapsed = 0
            if admission_date:
                try:
                    if isinstance(admission_date, str):
                        admission_dt = datetime.fromisoformat(admission_date.replace('Z', '+00:00'))
                    else:
                        admission_dt = admission_date
                    days_diff = (datetime.now() - admission_dt).days
                    days_elapsed = max(0, days_diff)
                except Exception:
                    days_elapsed = 0

            pipeline = [
                {'$match': {'patient_id': ObjectId(id)}},
                {'$group': {'_id': None, 'total_sales': {'$sum': '$amount'}}}
            ]
            canteen_result = list(mongo.db.canteen_sales.aggregate(pipeline))
            canteen_total = canteen_result[0]['total_sales'] if canteen_result else 0

            monthly_fee_raw = patient.get('monthlyFee', '0')
            monthly_fee = calculate_prorated_fee(monthly_fee_raw, days_elapsed)

            laundry_amount = patient.get('laundryAmount', 0) if patient.get('laundryStatus', False) else 0
            received_amount = int(str(patient.get('receivedAmount', '0')).replace(',', '') or '0')

            total_charges = monthly_fee + canteen_total + laundry_amount
            balance_due = total_charges - received_amount

            bill_data = {
                'Patient Name': patient.get('name', ''),
                'Father Name': patient.get('fatherName', ''),
                'CNIC': patient.get('cnic', ''),
                'Admission Date': patient.get('admissionDate', ''),
                'Discharge Date': patient.get('dischargeDate', '') or datetime.now().strftime('%Y-%m-%d'),
                'Days Stayed': days_elapsed,
                'Monthly Fee': monthly_fee,
                'Canteen Charges': canteen_total,
                'Laundry Charges': laundry_amount,
                'Total Charges': total_charges,
                'Amount Paid': received_amount,
                'Balance Due': balance_due
            }

            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df = pd.DataFrame([bill_data])
                df.to_excel(writer, index=False, sheet_name='Discharge Bill')

                worksheet = writer.sheets['Discharge Bill']
                worksheet.page_setup.paperSize = 9
                worksheet.page_setup.orientation = 'portrait'
                worksheet.page_setup.fitToPage = True
                worksheet.page_setup.fitToWidth = 1
                worksheet.page_setup.fitToHeight = 1
                worksheet.page_setup.scale = None
                worksheet.page_margins.left = 0.5
                worksheet.page_margins.right = 0.5
                worksheet.page_margins.top = 0.5
                worksheet.page_margins.bottom = 0.5
                worksheet.page_margins.header = 0.3
                worksheet.page_margins.footer = 0.3
                worksheet.print_options.horizontalCentered = True

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

                for cell in worksheet[1]:
                    cell.font = Font(bold=True, size=10, color='FFFFFF')
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = thin_border

                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    for cell in row:
                        cell.font = Font(size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        cell.border = thin_border
                        if 'Total' in str(worksheet.cell(1, cell.column).value) or 'Balance' in str(worksheet.cell(1, cell.column).value):
                            cell.fill = total_fill
                            cell.font = Font(size=10, bold=True)

                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    worksheet.column_dimensions[column_letter].width = min(max_length + 2, 25)

                worksheet.row_dimensions[1].height = 30
                for row_number in range(2, worksheet.max_row + 1):
                    worksheet.row_dimensions[row_number].height = 20

            output.seek(0)
            filename = f"discharge_bill_{patient.get('name', 'patient').replace(' ', '_')}.xlsx"
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        except Exception as error:
            print(f"Discharge Bill Error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/patients/<id>/payment_history', methods=['GET'])
    @role_required(['Admin'])
    def get_patient_payment_history(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            patient = mongo.db.patients.find_one({'_id': ObjectId(id)})
            if not patient:
                return jsonify([])

            target_name = patient.get('name', '').strip().lower()
            target_id_str = str(id)
            cursor = mongo.db.expenses.find({
                'type': 'incoming',
                'category': 'Patient Fee'
            }).sort('date', 1)

            history = []
            for doc in cursor:
                is_match = False
                doc_patient_id = str(doc.get('patient_id', ''))
                if doc_patient_id == target_id_str:
                    is_match = True

                if not is_match:
                    note = doc.get('note', '').lower()
                    if target_name and f"from {target_name}" in note:
                        is_match = True

                if is_match:
                    date_str = '-'
                    if doc.get('date'):
                        if isinstance(doc['date'], str):
                            date_str = doc['date'][:10]
                        else:
                            date_str = doc['date'].strftime('%d-%b-%Y')

                    history.append({
                        'date': date_str,
                        'amount': doc.get('amount', 0),
                        'method': doc.get('payment_method', 'Cash'),
                        'note': doc.get('note', '')
                    })
            return jsonify(history)
        except Exception as error:
            print(f"History error: {error}")
            return jsonify([])

    @app.route('/api/old-balances', methods=['GET'])
    @role_required(['Admin'])
    def get_old_balances():
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            month = request.args.get('month', type=int)
            year = request.args.get('year', type=int)
            query = {}
            if month and year:
                query = {
                    '$or': [
                        {'month': month, 'year': year},
                        {
                            'month': {'$exists': False},
                            'year': {'$exists': False},
                            'created_at': {
                                '$gte': datetime(year, month, 1),
                                '$lt': datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
                            }
                        }
                    ]
                }

            cursor = mongo.db.old_balances.find(query).sort('created_at', -1)
            balances = []
            for balance in cursor:
                balances.append({
                    'id': str(balance['_id']),
                    'name': balance.get('name', ''),
                    'amount': balance.get('amount', 0),
                    'commitment_date': balance.get('commitment_date', ''),
                    'last_call_date': balance.get('last_call_date', ''),
                    'note': balance.get('note', '')
                })
            return jsonify(balances)
        except Exception as error:
            print(f"Old Balance Fetch Error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/old-balances', methods=['POST'])
    @role_required(['Admin'])
    def add_old_balance():
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        data = clean_input_data(request.json)
        try:
            month = int(data.get('month') or datetime.now().month)
            year = int(data.get('year') or datetime.now().year)
            record = {
                'name': data.get('name'),
                'amount': int(data.get('amount', 0)),
                'commitment_date': data.get('commitment_date'),
                'last_call_date': data.get('last_call_date'),
                'note': data.get('note', ''),
                'month': month,
                'year': year,
                'created_at': datetime.now(),
                'added_by': session.get('username')
            }
            result = mongo.db.old_balances.insert_one(record)
            return jsonify({"message": "Record added", "id": str(result.inserted_id)}), 201
        except Exception as error:
            return jsonify({"error": str(error)}), 500

    @app.route('/api/old-balances/<id>', methods=['DELETE'])
    @role_required(['Admin'])
    def delete_old_balance(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            mongo.db.old_balances.delete_one({'_id': ObjectId(id)})
            return jsonify({"message": "Record deleted"})
        except Exception as error:
            return jsonify({"error": str(error)}), 500

    @app.route('/api/manual-discharge-receipts', methods=['GET'])
    @role_required(['Admin'])
    def list_manual_discharge_receipts():
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            query = (request.args.get('q') or '').strip().lower()
            cursor = mongo.db.manual_discharge_receipts.find().sort('created_at', -1)
            rows = []
            for doc in cursor:
                row = _serialize_manual_receipt(doc)
                if query:
                    haystack = f"{row.get('patient_name', '')} {row.get('father_name', '')} {row.get('contact_no', '')} {row.get('cnic', '')}".lower()
                    if query not in haystack:
                        continue
                rows.append(row)
            return jsonify(rows)
        except Exception as error:
            print(f"Manual receipt list error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/manual-discharge-receipts/<id>', methods=['GET'])
    @role_required(['Admin'])
    def get_manual_discharge_receipt(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            doc = mongo.db.manual_discharge_receipts.find_one({'_id': ObjectId(id)})
            if not doc:
                return jsonify({"error": "Record not found"}), 404
            return jsonify(_serialize_manual_receipt(doc))
        except Exception as error:
            print(f"Manual receipt get error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/manual-discharge-receipts', methods=['POST'])
    @role_required(['Admin'])
    def create_manual_discharge_receipt():
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            data = clean_input_data(request.json or {})
            now = datetime.now()

            fee_amount = _safe_int_amount(data.get('fee_amount'))
            rehab_next_month_amount = _safe_int_amount(data.get('rehab_next_month_amount'))
            test_amount = _safe_int_amount(data.get('test_amount'))
            canteen_amount = _safe_int_amount(data.get('canteen_amount'))
            laundry_amount = _safe_int_amount(data.get('laundry_amount'))
            barbar_amount = _safe_int_amount(data.get('barbar_amount'))
            medicine_amount = _safe_int_amount(data.get('medicine_amount'))
            other_amount = _safe_int_amount(data.get('other_amount'))
            received_amount = _safe_int_amount(data.get('received_amount'))
            gross_total = fee_amount + rehab_next_month_amount + test_amount + canteen_amount + laundry_amount + barbar_amount + medicine_amount + other_amount
            net_balance = gross_total - received_amount

            patient_id = data.get('patient_id') or ''
            if patient_id and ObjectId.is_valid(patient_id):
                patient_id = ObjectId(patient_id)
            else:
                patient_id = None

            payload = {
                'patient_id': patient_id,
                'patient_name': data.get('patient_name', ''),
                'father_name': data.get('father_name', ''),
                'age': data.get('age', ''),
                'cnic': data.get('cnic', ''),
                'contact_no': data.get('contact_no', ''),
                'area': data.get('area', ''),
                'address': data.get('address', ''),
                'admission_date': data.get('admission_date', ''),
                'discharge_date': data.get('discharge_date', ''),
                'stay_days': _safe_int_amount(data.get('stay_days')),
                'monthly_fee': _safe_int_amount(data.get('monthly_fee')),
                'fee_amount': fee_amount,
                'rehab_next_month_amount': rehab_next_month_amount,
                'test_amount': test_amount,
                'canteen_amount': canteen_amount,
                'laundry_amount': laundry_amount,
                'barbar_amount': barbar_amount,
                'medicine_amount': medicine_amount,
                'other_amount': other_amount,
                'received_amount': received_amount,
                'net_balance': net_balance,
                'notes': data.get('notes', ''),
                'created_by': session.get('username', 'Admin'),
                'updated_by': session.get('username', 'Admin'),
                'created_at': now,
                'updated_at': now
            }

            result = mongo.db.manual_discharge_receipts.insert_one(payload)
            return jsonify({"message": "Manual discharge receipt saved", "id": str(result.inserted_id)}), 201
        except Exception as error:
            print(f"Manual receipt create error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/manual-discharge-receipts/<id>', methods=['PUT'])
    @role_required(['Admin'])
    def update_manual_discharge_receipt(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            data = clean_input_data(request.json or {})
            fee_amount = _safe_int_amount(data.get('fee_amount'))
            rehab_next_month_amount = _safe_int_amount(data.get('rehab_next_month_amount'))
            test_amount = _safe_int_amount(data.get('test_amount'))
            canteen_amount = _safe_int_amount(data.get('canteen_amount'))
            laundry_amount = _safe_int_amount(data.get('laundry_amount'))
            barbar_amount = _safe_int_amount(data.get('barbar_amount'))
            medicine_amount = _safe_int_amount(data.get('medicine_amount'))
            other_amount = _safe_int_amount(data.get('other_amount'))
            received_amount = _safe_int_amount(data.get('received_amount'))
            gross_total = fee_amount + rehab_next_month_amount + test_amount + canteen_amount + laundry_amount + barbar_amount + medicine_amount + other_amount
            net_balance = gross_total - received_amount

            patient_id = data.get('patient_id') or ''
            if patient_id and ObjectId.is_valid(patient_id):
                patient_id = ObjectId(patient_id)
            else:
                patient_id = None

            payload = {
                'patient_id': patient_id,
                'patient_name': data.get('patient_name', ''),
                'father_name': data.get('father_name', ''),
                'age': data.get('age', ''),
                'cnic': data.get('cnic', ''),
                'contact_no': data.get('contact_no', ''),
                'area': data.get('area', ''),
                'address': data.get('address', ''),
                'admission_date': data.get('admission_date', ''),
                'discharge_date': data.get('discharge_date', ''),
                'stay_days': _safe_int_amount(data.get('stay_days')),
                'monthly_fee': _safe_int_amount(data.get('monthly_fee')),
                'fee_amount': fee_amount,
                'rehab_next_month_amount': rehab_next_month_amount,
                'test_amount': test_amount,
                'canteen_amount': canteen_amount,
                'laundry_amount': laundry_amount,
                'barbar_amount': barbar_amount,
                'medicine_amount': medicine_amount,
                'other_amount': other_amount,
                'received_amount': received_amount,
                'net_balance': net_balance,
                'notes': data.get('notes', ''),
                'updated_by': session.get('username', 'Admin'),
                'updated_at': datetime.now()
            }

            result = mongo.db.manual_discharge_receipts.update_one(
                {'_id': ObjectId(id)},
                {'$set': payload}
            )
            if result.matched_count == 0:
                return jsonify({"error": "Record not found"}), 404
            return jsonify({"message": "Manual discharge receipt updated"})
        except Exception as error:
            print(f"Manual receipt update error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/manual-discharge-receipts/<id>', methods=['DELETE'])
    @role_required(['Admin'])
    def delete_manual_discharge_receipt(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            result = mongo.db.manual_discharge_receipts.delete_one({'_id': ObjectId(id)})
            if result.deleted_count == 0:
                return jsonify({"error": "Record not found"}), 404
            return jsonify({"message": "Record deleted"})
        except Exception as error:
            print(f"Manual receipt delete error: {error}")
            return jsonify({"error": str(error)}), 500

    @app.route('/api/patients/<id>/bill/preview', methods=['GET'])
    @role_required(['Admin', 'Family'])
    def preview_patient_bill(id):
        if not check_db():
            return jsonify({"error": "Database error"}), 500

        try:
            from services.pdf_engine import generate_billing_pdf
            from services.site_profile import get_site_profile

            patient = mongo.db.patients.find_one({'_id': ObjectId(id)})
            if not patient:
                return "Patient not found", 404

            user_id = get_current_user_id()
            user = mongo.db.users.find_one({'_id': ObjectId(user_id)})
            if user and user.get('role') == 'Family':
                if ObjectId(id) not in user.get('patient_ids', []):
                    return "Unauthorized: You do not have access to this patient's bill", 403

            canteen_agg = list(mongo.db.canteen_sales.aggregate([
                {'$match': {'patient_id': ObjectId(id)}},
                {'$group': {'_id': None, 'total': {'$sum': '$amount'}}}
            ]))
            canteen_total = canteen_agg[0]['total'] if canteen_agg else 0
            financial = patient_financial_summary(patient, canteen_total, month_year=datetime.now().strftime('%B %Y'))

            patient_data = {
                '_id': str(patient['_id']),
                'name': decrypt_data(patient.get('name', '')),
                'fatherName': patient.get('fatherName', ''),
                'cnic': decrypt_data(patient.get('cnic', '')),
                'contactNo': decrypt_data(patient.get('contactNo', '')),
                'address': patient.get('address', ''),
                'admissionDate': str(patient.get('admissionDate', '')),
            }

            pdf_bytes, err = generate_billing_pdf(patient_data, financial, get_site_profile(mongo))
            if err:
                response = make_response(pdf_bytes)
                response.headers['Content-Type'] = 'text/html'
                return response

            response = make_response(pdf_bytes)
            response.headers['Content-Type'] = 'application/pdf'
            response.headers['Content-Disposition'] = f'inline; filename=bill_{id}.pdf'
            return response
        except Exception as error:
            print(f"Preview Error: {error}")
            return str(error), 500
